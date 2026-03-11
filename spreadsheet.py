"""
spreadsheet.py
Builds a formatted Excel workbook from a list of extracted payslip dicts.
"""

from datetime import datetime

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from config import FIELDS


# Module-level style constants — openpyxl copies these per-cell internally,
# so creating them once here is safe and avoids instantiating identical objects
# for every header cell.
_HEADER_FILL      = PatternFill(fill_type="solid", fgColor="D9E1F2")  # blue-grey — canonical
_EXTRA_HEADER_FILL = PatternFill(fill_type="solid", fgColor="FFD966")  # amber — dynamic fields
_BOLD = Font(bold=True)
_CURRENCY_FORMAT = '£#,##0.00'

# Field keys whose values are £ amounts — these receive the currency number format.
# All fields from basic_pay onwards in config.FIELDS are monetary.
# Text/label fields (provider, names, period_label, period_date, tax_code,
# ni_number) are deliberately excluded.
_CURRENCY_KEYS = {
    "basic_pay", "pension_employee", "student_loan", "ni_employee",
    "paye_tax", "total_deductions", "take_home_pay", "ni_employer",
    "pension_employer", "ytd_gross", "ytd_taxable", "ytd_tax_paid",
    "ytd_ni_employee", "ytd_pension_employee", "ytd_pension_employer",
}


def build_workbook(records: list[dict]) -> openpyxl.Workbook:
    """
    Build a formatted Excel workbook from a list of extracted payslip dicts.

    Args:
        records: List of dicts, one per payslip, with keys matching config.FIELDS.
                 None values are written as empty cells.

    Returns:
        openpyxl Workbook with a single "Payslips" sheet. Does not save to disk.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payslips"

    # Only include canonical columns that have at least one non-None value
    # across all records. This suppresses provider-specific excluded fields
    # (e.g. Capium's excluded YTD columns) without needing special-case logic.
    canonical_keys = [
        key for key, _ in FIELDS
        if any(record.get(key) is not None for record in records)
    ]
    canonical_names = [name for key, name in FIELDS if key in canonical_keys]
    canonical_set = set(key for key, _ in FIELDS)

    # Collect extra (dynamic) keys in first-seen order across all records.
    seen_extra: set[str] = set()
    extra_keys: list[str] = []
    for record in records:
        for k in record:
            if k not in canonical_set and k not in seen_extra:
                extra_keys.append(k)
                seen_extra.add(k)

    extra_names = [k.replace('_', ' ').title() for k in extra_keys]

    all_keys = canonical_keys + extra_keys
    all_names = canonical_names + extra_names
    n_canonical = len(canonical_keys)

    # --- Header row (row 1) ---
    for col_idx, name in enumerate(all_names, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = _BOLD
        if col_idx > n_canonical:
            cell.fill = _EXTRA_HEADER_FILL
            note = Comment(
                "Unrecognised field detected automatically.\n"
                "Review and add to the schema in the next iteration if needed.",
                "Payslip Collator",
            )
            note.width = 240
            note.height = 60
            cell.comment = note
        else:
            cell.fill = _HEADER_FILL

    # --- Sort records by period_date ascending ---
    # period_date is stored as "DD/MM/YYYY". Rows with None or unparseable
    # dates use datetime.max as a sentinel so they sink to the bottom rather
    # than raising an exception.
    def _sort_key(record: dict) -> datetime:
        pd = record.get("period_date")
        if not pd:
            return datetime.max
        try:
            return datetime.strptime(pd, "%d/%m/%Y")
        except ValueError:
            return datetime.max

    sorted_records = sorted(records, key=_sort_key)

    # --- Data rows (rows 2+) ---
    for row_idx, record in enumerate(sorted_records, start=2):
        for col_idx, key in enumerate(all_keys, start=1):
            value = record.get(key)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Extra fields are always monetary; canonical currency keys are listed explicitly.
            if (key in _CURRENCY_KEYS or key in seen_extra) and value is not None:
                cell.number_format = _CURRENCY_FORMAT

    # --- Auto-fit column widths ---
    # openpyxl has no built-in auto-fit, so we measure the longest content
    # in each column. For currency cells we simulate the display string
    # (e.g. "£1,234.56") because that is what determines visible width —
    # the raw float "1234.56" would underestimate. Capped at 50 to prevent
    # very long employer names from making columns unusably wide.
    for col_idx, (key, display_name) in enumerate(zip(all_keys, all_names), start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(display_name)
        is_currency = key in _CURRENCY_KEYS or key in seen_extra

        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            if is_currency:
                display_str = f"£{value:,.2f}"
            else:
                display_str = str(value)
            max_len = max(max_len, len(display_str))

        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    return wb
